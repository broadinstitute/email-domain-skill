"""Reading and writing Excel workbooks on local disk, via openpyxl.

This is the whole of the spreadsheet layer. It replaces the Sheets API wrapper, along with the
hand-rolled A1 helpers that went with it — openpyxl supplies `get_column_letter` and
`cell.coordinate`.

Cells are read with `data_only=True`, so a formula yields the value Excel last cached for it.
A workbook written by a tool that does not compute formulas has no cached value, and such a
cell reads as empty; that is noted in the README's limitations.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class Cell:
    """One non-empty cell, with its text as a reader of the workbook would see it."""

    sheet_title: str
    row: int  # 1-based, as in A1 notation
    column: int  # 1-based
    text: str

    @property
    def a1(self) -> str:
        return f'{get_column_letter(self.column)}{self.row}'


def sheet_titles(path: Path) -> list[str]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def read_cells(path: Path) -> list[Cell]:
    """Every non-empty cell of every sheet, in reading order."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        cells: list[Cell] = []
        for title in workbook.sheetnames:
            worksheet = workbook[title]
            for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                for column_index, value in enumerate(row, start=1):
                    text = _text(value)
                    if text:
                        cells.append(Cell(title, row_index, column_index, text))
        return cells
    finally:
        workbook.close()


def _text(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, bool):
        return ''  # Never a domain, and str(True) would only add noise to scans.
    return str(value)


def create(path: Path, sheets: dict[str, list[list[str]]]) -> None:
    """Write a new workbook with one sheet per entry, each seeded with the given rows."""
    from openpyxl import Workbook

    workbook = Workbook()
    default = workbook.active
    if default is not None:
        workbook.remove(default)
    for title, rows in sheets.items():
        worksheet = workbook.create_sheet(title)
        for row in rows:
            worksheet.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def read_rows(path: Path, sheet_title: str) -> list[list[str]]:
    """Every row of one sheet as strings, trailing empty cells preserved as ''."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_title not in workbook.sheetnames:
            return []
        worksheet = workbook[sheet_title]
        return [
            ['' if value is None else str(value).strip() for value in row]
            for row in worksheet.iter_rows(values_only=True)
        ]
    finally:
        workbook.close()


def write_cells(path: Path, values: dict[str, list[tuple[int, int, str]]]) -> int:
    """Set individual cells, given `{sheet_title: [(row, column, text), ...]}`.

    Cell by cell rather than by rectangle: only the cells named here are touched, so a redaction
    can never write over a neighbouring cell that nobody approved changing. Returns how many were
    written.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    try:
        written = 0
        for title, cells in values.items():
            if title not in workbook.sheetnames:
                raise KeyError(f'{path} has no sheet named {title!r}.')
            worksheet = workbook[title]
            for row, column, text in cells:
                worksheet.cell(row=row, column=column, value=text)
                written += 1
        workbook.save(path)
        return written
    finally:
        workbook.close()


def rewrite(path: Path, sheets: dict[str, list[list[str]]]) -> None:
    """Replace the contents of the named sheets, leaving other sheets alone.

    Used only for the analysis workbook, which this server owns outright. A metrics workbook is
    never rewritten wholesale — redaction goes through `write_cells`, which touches only the cells
    it was asked to.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    try:
        for title, rows in sheets.items():
            # Re-create in place rather than appending: clearing a sheet row by row is slower,
            # and dropping it would silently reorder the workbook's tabs on every write.
            index = workbook.sheetnames.index(title) if title in workbook.sheetnames else None
            if index is not None:
                workbook.remove(workbook[title])
            worksheet = workbook.create_sheet(title, index)
            for row in rows:
                worksheet.append(row)
        workbook.save(path)
    finally:
        workbook.close()
