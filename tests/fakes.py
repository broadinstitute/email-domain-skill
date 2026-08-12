"""In-memory stand-in for the Google Drive MCP connector.

Much smaller than the Sheets fake it replaces, because cell semantics are no longer faked:
workbooks are real `.xlsx` files built with openpyxl, so the tests exercise the same reader and
writer that production uses. This only has to model Drive's file-level behaviour — ids, names,
mime types, `modifiedTime`, and the fact that `create` always makes a *new* file, since the
connector offers no update.
"""

from __future__ import annotations

import itertools
from dataclasses import dataclass, field
from pathlib import Path

from email_domain_scrubber import xlsx
from email_domain_scrubber.drive import XLSX_MIME, FileInfo
from email_domain_scrubber.errors import WorkbookNotFound


def write_xlsx(path: Path, sheets: dict[str, list[list[str]]]) -> Path:
    """Build a real `.xlsx` on disk. The fixture workbook for most tests."""
    xlsx.create(path, sheets)
    return path


@dataclass
class FakeFile:
    file_id: str
    name: str
    content: bytes
    mime_type: str = XLSX_MIME
    modified_time: str = '2026-01-01T00:00:00Z'
    parent_id: str = ''

    def info(self) -> FileInfo:
        return FileInfo(
            file_id=self.file_id,
            name=self.name,
            mime_type=self.mime_type,
            modified_time=self.modified_time,
            parent_id=self.parent_id,
        )


class FakeDrive:
    """A `Drive` over a dictionary."""

    def __init__(self) -> None:
        self.files: dict[str, FakeFile] = {}
        self.created: list[FakeFile] = []
        self.downloads: list[str] = []
        self._ids = itertools.count(1)

    # -- test helpers --------------------------------------------------------------------
    def add_workbook(
        self,
        name: str,
        sheets: dict[str, list[list[str]]],
        tmp_path: Path,
        *,
        parent: str | None = None,
    ) -> FakeFile:
        """Register an `.xlsx` in fake Drive, built from `sheets`."""
        source = write_xlsx(tmp_path / f'source-{next(self._ids)}.xlsx', sheets)
        return self.add_bytes(name, source.read_bytes(), parent=parent)

    def add_bytes(
        self, name: str, content: bytes, *, mime_type: str = XLSX_MIME, parent: str | None = None
    ) -> FakeFile:
        file = FakeFile(
            file_id=f'file{next(self._ids):017d}',
            name=name,
            content=content,
            mime_type=mime_type,
            parent_id=parent or '',
        )
        self.files[file.file_id] = file
        return file

    def touch(self, file_id: str, content: bytes, modified_time: str) -> None:
        """Replace a file's bytes and bump its `modifiedTime`, as an editor would."""
        file = self._file(file_id)
        file.content = content
        file.modified_time = modified_time

    def _file(self, file_id: str) -> FakeFile:
        try:
            return self.files[file_id]
        except KeyError:
            raise WorkbookNotFound(f'no such file: {file_id}') from None

    # -- Drive ---------------------------------------------------------------------------
    async def get_metadata(self, file_id: str) -> FileInfo:
        return self._file(file_id).info()

    async def download(self, file_id: str) -> bytes:
        self.downloads.append(file_id)
        return self._file(file_id).content

    async def create(
        self, name: str, content: bytes, *, mime_type: str = XLSX_MIME, parent_id: str | None = None
    ) -> FileInfo:
        file = self.add_bytes(name, content, mime_type=mime_type, parent=parent_id)
        self.created.append(file)
        return file.info()

    async def search(self, query: str, *, page_size: int = 10) -> list[FileInfo]:
        return [file.info() for file in list(self.files.values())[:page_size]]


@dataclass
class Recorder:
    """Captures what the Excel MCP server would have been asked to do.

    Redaction writes are performed by an external server in production. Tests apply the same
    blocks with openpyxl, which is what that server does internally, and keep the calls so a
    test can assert on their shape.
    """

    calls: list[tuple[str, str, str, list[list[str]]]] = field(default_factory=list)

    def apply(self, path: Path, blocks: list) -> None:
        """Stand in for `write_data_to_excel`, one call per block."""
        from openpyxl import load_workbook
        from openpyxl.utils.cell import coordinate_to_tuple

        for block in blocks:
            self.calls.append((str(path), block.sheet, block.start_cell, block.values))
            row, column = coordinate_to_tuple(block.start_cell)
            workbook = load_workbook(path)
            try:
                worksheet = workbook[block.sheet]
                for row_offset, values in enumerate(block.values):
                    for column_offset, value in enumerate(values):
                        worksheet.cell(
                            row=row + row_offset, column=column + column_offset, value=value
                        )
                workbook.save(path)
            finally:
                workbook.close()
