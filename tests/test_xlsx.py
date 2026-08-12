"""The local Excel layer that replaced the Sheets API wrapper."""

from openpyxl import load_workbook

from email_domain_scrubber import xlsx


def test_read_cells_reports_a1_addresses_and_skips_blanks(tmp_path):
    path = tmp_path / 'w.xlsx'
    xlsx.create(path, {'Users': [['User', 'Email'], ['Alice', None], [None, 'a@lab.io']]})

    cells = {(cell.sheet_title, cell.a1): cell.text for cell in xlsx.read_cells(path)}
    assert cells == {
        ('Users', 'A1'): 'User',
        ('Users', 'B1'): 'Email',
        ('Users', 'A2'): 'Alice',
        ('Users', 'B3'): 'a@lab.io',
    }


def test_read_cells_covers_every_sheet(tmp_path):
    path = tmp_path / 'w.xlsx'
    xlsx.create(path, {'One': [['a@one.org']], 'Two': [['b@two.org']]})

    assert {cell.sheet_title for cell in xlsx.read_cells(path)} == {'One', 'Two'}


def test_read_cells_handles_columns_past_z(tmp_path):
    """Column 27 is AA, which a naive A-plus-index scheme gets wrong."""
    path = tmp_path / 'w.xlsx'
    xlsx.create(path, {'S': [[None] * 26 + ['pi@lab.io']]})

    assert [cell.a1 for cell in xlsx.read_cells(path)] == ['AA1']


def test_numbers_become_text_and_booleans_are_ignored(tmp_path):
    path = tmp_path / 'w.xlsx'
    xlsx.create(path, {'S': [[12, True, 'a@lab.io']]})

    assert [cell.text for cell in xlsx.read_cells(path)] == ['12', 'a@lab.io']


def test_rewrite_replaces_a_sheet_without_reordering_the_workbook(tmp_path):
    path = tmp_path / 'w.xlsx'
    xlsx.create(path, {'First': [['1']], 'Target': [['old']], 'Last': [['3']]})

    xlsx.rewrite(path, {'Target': [['new', 'row']]})

    assert xlsx.sheet_titles(path) == ['First', 'Target', 'Last']
    assert xlsx.read_rows(path, 'Target') == [['new', 'row']]
    assert xlsx.read_rows(path, 'First') == [['1']]


def test_rewrite_leaves_no_stale_rows_behind(tmp_path):
    path = tmp_path / 'w.xlsx'
    xlsx.create(path, {'S': [['a'], ['b'], ['c']]})

    xlsx.rewrite(path, {'S': [['only']]})

    assert xlsx.read_rows(path, 'S') == [['only']]


def test_created_workbook_has_no_leftover_default_sheet(tmp_path):
    path = tmp_path / 'w.xlsx'
    xlsx.create(path, {'Only': [['x']]})

    assert load_workbook(path).sheetnames == ['Only']


def test_read_rows_of_an_absent_sheet_is_empty(tmp_path):
    path = tmp_path / 'w.xlsx'
    xlsx.create(path, {'S': [['x']]})

    assert xlsx.read_rows(path, 'Missing') == []
